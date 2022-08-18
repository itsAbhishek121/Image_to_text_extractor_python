# Program to Extract Metadata(Title of book, Name of authors, ISBN number and Publisher name) 
# of book cover pages using Tesseract OCR.

from PIL import Image
import pytesseract as pt
from pytesseract import Output
import os
from difflib import SequenceMatcher
import xlsxwriter
import spacy
import re
from openpyxl import load_workbook
import cv2
import sys

def single_image_ocr(path):
    pt.pytesseract.tesseract_cmd = 'C:\\Users\\Asus\\AppData\\Local\\Programs\\Tesseract-OCR\\tesseract.exe'
    try:
        image = cv2.imread(path)
        gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)   # Converting image into gray scale image
        img = cv2.threshold(gray_image, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]   # Converting it to binary image by Thresholding
        # Applying Tesseract OCR
        text = pt.image_to_string(img, lang ="eng", config='--psm 11')
        tx = pt.image_to_data(img, lang="eng",output_type=Output.DICT,config='--psm 11')
        n_boxes = len(tx['level'])
        max_height = 0
        corr_block = 0
        title = ""
        for i in range(n_boxes):
            if ((tx['text'][i] != '') and (tx['text'][i] != ' ') and (tx['height'][i] > max_height)):
                max_height = tx['height'][i]
                corr_block = tx['block_num'][i]
        for i in range(n_boxes):
            if ((tx['block_num'][i] == corr_block) and (tx['text'][i] != '')):
                title = title + tx['text'][i] + " "
        
        nlp = spacy.load('en_core_web_lg')
        doc = nlp(text)
        # Listing entities that falls under PERSON and ORG entity
        lst = []
        lst1 = []
        for ent in doc.ents:
            if (ent.label_ == "PERSON"):
                lst.append(ent.text)
            if (ent.label_ == "ORG"):
                lst1.append(ent.text)
        # Storing each line of recognized as a single string
        l = []
        s= ""
        for i in text:
            if (i == "\n"):
                l.append(s)
                s = ""
            else:
                s = s + i
        # Ensuring correctness of author name by matching actual text recognized and the entity NER(Named Entity Recognization) identified.
        extra = []
        for i in lst:
            x = 0
            temp = ""
            for j in l:
                if (SequenceMatcher(None,i,j).ratio() > x):
                    x = SequenceMatcher(None,i,j).ratio()
                    temp = j
            extra.append(temp)
        extra = list(dict.fromkeys(extra))
        author = ""
        for i in range(len(extra)):
            if (i==len(extra)-1):
                author = author + extra[i]
            else:
                author = author + extra[i] + ", "
        
        extra1 = []
        for i in lst1:
            x = 0
            temp = ""
            for j in l:
                if (SequenceMatcher(None,i,j).ratio() > x):
                    x = SequenceMatcher(None,i,j).ratio()
                    temp = j
            extra1.append(temp)
        extra1 = list(dict.fromkeys(extra1))
        publisher = ""
        for i in range(len(extra1)):
            if (i==len(extra1)-1):
                publisher = publisher + extra1[i]
            else:
                publisher = publisher + extra1[i] + ", "
        temp = []
        isbn = ""
        for i in range(len(text)):
            if (text[i:i+4].upper() == 'ISBN'):
                result = re.search("[^\n]*", text[i::])
                isbn = result.group()
                temp.append(isbn)
        x = ", ".join(temp)
        # Creating a Workbook
        workbook = xlsxwriter.Workbook('output_single.xlsx')
        # Adding new Worksheet
        worksheet = workbook.add_worksheet()

        # Naming the Headers
        worksheet.write('A1', ' Title of the Book ')
        worksheet.write('B1', ' Names of the Authors ')
        worksheet.write('C1', ' ISBN Numbers')
        worksheet.write('D1', ' Publishers ')
        
        # Feeding data into XLSX file.
        worksheet.write(1, 0, title)
        worksheet.write(1, 1, author)
        worksheet.write(1, 2, isbn)
        worksheet.write(1, 3, publisher)

        # Closing Workbook after writing output in XLSX file
        workbook.close()

        wrkbk = load_workbook("C:\\Users\\Asus\\Desktop\\305\\output_single.xlsx")
        sh = wrkbk.active
        rows_aff = sh.max_row
        return rows_aff
    except Exception as e:
        print("1. Error Occured: ",e.__class__)

def multiple_image_ocr(path):
    text_list,tx_list,n = using_ocr_on_images(path)

    title_list =  finding_titles(tx_list)
    author_list = finding_author_names(text_list)
    isbn_list = finding_isbn(text_list)
    publisher_list = finding_publisher(text_list)

    rows_aff = storing_output_in_XLSX(n,title_list,author_list,isbn_list,publisher_list)
    return rows_aff

def using_ocr_on_images(path):
    pt.pytesseract.tesseract_cmd = 'C:\\Users\\Asus\\AppData\\Local\\Programs\\Tesseract-OCR\\tesseract.exe'
    text_list = []
    tx_list = []
    n = 0
    try:
        for imageName in os.listdir(path):
            inputPath = os.path.join(path, imageName)
            image = cv2.imread(inputPath)
            gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)   # Converting image into gray scale image
            img = cv2.threshold(gray_image, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]   # Converting it to binary image by Thresholding

            # Applying Tesseract OCR
            text = pt.image_to_string(img, lang ="eng", config='--psm 11')
            tx = pt.image_to_data(img, lang="eng",output_type=Output.DICT,config='--psm 11')
            text_list.append(text)
            tx_list.append(tx)
            n+=1
    except Exception as e:
        print("2. Error Occured: ",e.__class__)
    return text_list,tx_list,n

def finding_titles(tx_list):
    title_list = []
    try:
        for tx in tx_list:
            n_boxes = len(tx['level'])
            max_height = 0
            corr_block = 0
            title = ""
            for i in range(n_boxes):
                if ((tx['text'][i] != '') and (tx['text'][i] != ' ') and (tx['height'][i] > max_height)):
                    max_height = tx['height'][i]
                    corr_block = tx['block_num'][i]
            for i in range(n_boxes):
                if ((tx['block_num'][i] == corr_block) and (tx['text'][i] != '')):
                    title = title + tx['text'][i] + " "
            title_list.append(title)
    except Exception as e:
        print("3. Error Occured: ",e.__class__)
    return title_list

def finding_author_names(text_list):
    nlp = spacy.load('en_core_web_lg')
    author_list = []
    try:
        for text in text_list:
            doc = nlp(text)
            # Listing entities that falls under PERSON entity
            lst = []
            for ent in doc.ents:
                if (ent.label_ == "PERSON"):
                    lst.append(ent.text)
            # Storing each line of recognized as a single string
            l = []
            s= ""
            for i in text:
                if (i == "\n"):
                    l.append(s)
                    s = ""
                else:
                    s = s + i
            # Ensuring correctness of author name by matching actual text recognized and the entity NER(Named Entity Recognization) identified.
            extra = []
            for i in lst:
                x = 0
                temp = ""
                for j in l:
                    if (SequenceMatcher(None,i,j).ratio() > x):
                        x = SequenceMatcher(None,i,j).ratio()
                        temp = j
                extra.append(temp)
            extra = list(dict.fromkeys(extra))
            author = ""
            for i in range(len(extra)):
                if (i==len(extra)-1):
                    author = author + extra[i]
                else:
                    author = author + extra[i] + ", "
            author_list.append(author)
    except Exception as e:
        print("4. Error Occured: ",e.__class__)
    return author_list

def finding_isbn(text_list):
    isbn_list = []
    try:
        for text in text_list:
            temp = []
            isbn = ""
            for i in range(len(text)):
                if (text[i:i+4].upper() == 'ISBN'):
                    result = re.search("[^\n]*", text[i::])
                    isbn = result.group()
                    temp.append(isbn)
            x = ", ".join(temp)
            isbn_list.append(x)
    except Exception as e:
        print("5. Error Occured: ",e.__class__)
    return isbn_list
        
def finding_publisher(text_list):
    nlp = spacy.load('en_core_web_lg')
    publisher_list = []
    try:
        for text in text_list:
            doc = nlp(text)
            # Listing entities that falls under ORG entity
            lst = []
            for ent in doc.ents:
                if (ent.label_ == "ORG"):
                    lst.append(ent.text)
            # Storing each line of recognized as a single string
            l = []
            s= ""
            for i in text:
                if (i == "\n"):
                    l.append(s)
                    s = ""
                else:
                    s = s + i
            # Ensuring correctness of publisher name by matching actual text recognized and the entity NER(Named Entity Recognization) identified.
            extra = []
            for i in lst:
                x = 0
                temp = ""
                for j in l:
                    if (SequenceMatcher(None,i,j).ratio() > x):
                        x = SequenceMatcher(None,i,j).ratio()
                        temp = j
                extra.append(temp)
            extra = list(dict.fromkeys(extra))
            publisher = ""
            for i in range(len(extra)):
                if (i==len(extra)-1):
                    publisher = publisher + extra[i]
                else:
                    publisher = publisher + extra[i] + ", "
            publisher_list.append(publisher)
    except Exception as e:
        print("6. Error Occured: ",e.__class__)
    return publisher_list

def storing_output_in_XLSX(n,title_list,author_list,isbn_list,publisher_list):
    try:
        # Creating a Workbook
        workbook = xlsxwriter.Workbook('output.xlsx')
        # Adding new Worksheet
        worksheet = workbook.add_worksheet()

        # Naming the Headers
        worksheet.write('A1', ' Title of the Book ')
        worksheet.write('B1', ' Names of the Authors ')
        worksheet.write('C1', ' ISBN Numbers')
        worksheet.write('D1', ' Publishers ')
        
        # Feeding data into XLSX file.
        for i in range(n):
            worksheet.write(i+1, 0, title_list[i])
            worksheet.write(i+1, 1, author_list[i])
            worksheet.write(i+1, 2, isbn_list[i])
            worksheet.write(i+1, 3, publisher_list[i])

        # Closing Workbook after writing output in XLSX file
        workbook.close()

        wrkbk = load_workbook("C:\\Users\\Asus\\Desktop\\305\\output.xlsx")
        sh = wrkbk.active
        rows_aff = sh.max_row
        return rows_aff
    except Exception as e:
        print("7. Error Occured: ",e.__class__)

if __name__ == '__main__':
    flag = sys.argv[1]
    path = sys.argv[2]
    if (flag == 0):
        single_image_ocr(path)
    else :
        multiple_image_ocr(path)

